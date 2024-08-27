''' <---|---|---| IMPORTAÇÃO DE LIBS |---|---|---> '''

# importando o Tkinter
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog

# importando o PyPDF2
import PyPDF2

# importando o RegEx
import re

# importando meu dicionario de exemplos
import dic_faturas

# importando o google generativeai
import google.generativeai as genai

# importando o google api_core
import google.api_core
import google.api_core.exceptions

# importando o OpenPyXl
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# importando o datetime
import datetime

# importando o requests
import requests

# importando o pillow
from PIL import Image
from PIL import ImageTk

# importando o os
import os

# importando o locale
import locale

''' <---|---|---| IMPORTAÇÃO DE LIBS |---|---|---> '''





''' <---|---|---| TESTE DE CONEXÃO |---|---|---> '''

# função para verificar se a conexão com a internet foi estabelecida
def teste_conexao():
    # janela de erro
    janela_erro_conexao = Toplevel()
    janela_erro_conexao.geometry("400x250")
    janela_erro_conexao.title("PDF to Excel - Erro de Conexão")
    janela_erro_conexao.config(bg="#CDDDDD", pady=30, padx=30)
    
    erro_conexao_titulo = Label(janela_erro_conexao, text="Não foi possivel conectar a internet.",fg="red", bg="#CDDDDD", font=("Arial bold", 15))
    erro_conexao_titulo.pack(side="top", padx=5, pady=5)
    
    erro_conexao_label = Label(janela_erro_conexao, text="Ocorreu um erro ao tentar conectar com a internet.\nVerifique sua conexão e tente novamente", bg="#CDDDDD", anchor="w")
    erro_conexao_label.pack(side="top", padx=5, pady=5)
    
    ok_erro_conexao_btn = Button(janela_erro_conexao, text="Tentar Novamente", command=lambda: [janela_erro_conexao.destroy(), teste_conexao()])
    ok_erro_conexao_btn.pack(side="top", padx=5, pady=5)

    # verifica a conexão
    try:
        response = requests.get("https://www.google.com")
        if response.status_code == 200:
            janela_erro_conexao.destroy()
            
            return True  
        else:
            janela_erro_conexao.mainloop()
            
            return False  
    except requests.exceptions.RequestException:
            janela_erro_conexao.mainloop()
        
            return False  

''' <---|---|---| TESTE DE CONEXÃO |---|---|---> '''





''' <---|---|---| JANELA DE ERRO |---|---|---> '''

# função que cria uma janela de erro
def raise_janela_erro(erro_txt="", erro_msg="", erro_name_txt=""):
    # janela de erro
    error_window = Toplevel()
    error_window.geometry("400x250")
    error_window.title("PDF to Excel - Erro")
    error_window.config(bg="#CDDDDD", pady=30, padx=30)
    
    erro_titulo = Label(error_window, text="ERRO!",fg="red", bg="#CDDDDD", font=("Arial bold", 15))
    erro_titulo.pack(side="top", padx=5, pady=5)
    
    erro_sub = Label(error_window, text=erro_txt, bg="#CDDDDD", font=("Arial bold", 12))
    erro_sub.pack(side="top", padx=5, pady=2)
    
    if erro_msg != "":
        erro_sub_msg = Label(error_window, text=erro_msg, bg="#CDDDDD", font=("Arial bold", 10))
        erro_sub_msg.pack(side="top", padx=5, pady=2)
    
    if erro_name_txt != "":
        erro_name = Label(error_window, text=erro_name_txt, bg="#CDDDDD", fg="blue", font=("Arial", 10, "underline"), wraplength=300)
        erro_name.pack(side="top", padx=5, pady=2)
    
    ok_erro_btn = Button(error_window, text="Fechar", command=lambda: error_window.destroy())
    ok_erro_btn.pack(side="bottom", padx=5, pady=5)
    
    erro_msg = Label(error_window, text="Tente novamente mais tarde.", bg="#CDDDDD")
    erro_msg.pack(side="bottom", padx=5, pady=5)
    
    error_window.mainloop()
    
''' ERRO CASO NENHUM ARQUIVO PDF SEJA SELECIONADO '''
class ErroSemArquivosPDF(Exception):
    pass

''' ERRO CASO A CHAVE API NÃO TENHA SIDO ENVIADA '''
class ErroSemChaveAPI(Exception):
    pass

''' <---|---|---| JANELA DE ERRO |---|---|---> '''





''' <---|---|---| MANIPULAÇÃO DE ARQUIVOS |---|---|---> '''

# arquivos convertidos
pdfs = []

# função para selecionar e converter os arquivos
def procura_arquivos():
    # seletor de arquivos
    files = filedialog.askopenfilenames(initialdir = "/", title = "Selecione o arquivo", filetypes = (("PDF files", "*.pdf*"), ("all files", "*.*")))
    
    # convertendo os arquivos
    for file in files:
        pdfs.append(PyPDF2.PdfReader(open(file, "rb")))
    
    # mostra a mensagem na tela
    num_file = 1
    file_str = ""
    for x in files:
        nome_cortado = re.search(r".*/([^/]+)$", x)
        if num_file < 5:
            if len(nome_cortado.group(1)) <= 30:
                file_str += f"[{nome_cortado.group(1)}],\n"
            
            else:
                file_str += f"[{nome_cortado.group(1)[:31]}],\n"
                
        elif num_file == 5:
            if len(nome_cortado.group(1)) <= 30:
                file_str += f"[{nome_cortado.group(1)}]..."
             
            else:
                file_str += f"[{nome_cortado.group(1)[:31]}]..."
            break
        
        num_file += 1
    
    quant_arquivos.configure(text=f"Arquivos selecionados:\t{len(pdfs)}")
    nome_arquivos.configure(text=file_str)
    procura_arquivo_btn.configure(text="Selecionar mais arquivos", width=20)
    
    troca_arquivo_btn.pack(side="top", padx=5, pady=5)
  
# função para trocar os arquivos selecionados
def troca_arquivos():
    pdfs.clear()
    procura_arquivos()

  
# função que retorna os itens do dicionario de exemplos
def ex_files():
    dic_str = "Analize estes arquivos:"
    for dic_key, dic_obj in dic_faturas.ex_faturas.items():
        dic_str += f"\n\n{dic_key}:\n"
        
        for x, y in dic_obj.items():
            dic_str += f"{x}:\n{y}\n\n"
    
    return dic_str
    
# função para adicionar os arquivos ao prompt 
def prompt_files(files):
    num = 1
    prompt_str = ""
    
    for pdf in files:
        pdf_str = ""
        for n in range(len(pdf.pages)):
            pdf_str += pdf.pages[n].extract_text()
        prompt_str += f"pdf_{num}:\ndoc:\n{pdf_str}\n\nNumero fatura:\n\n\nImportador:\n\n\n"
        
        num += 1
        
    return prompt_str

''' <---|---|---| MANIPULAÇÃO DE ARQUIVOS |---|---|---> '''





''' <---|---|---| ARMAZENAMENTO DA CHAVE API E ESCOLHA DE USO DOS ARQUIVOS DE EXEMPLO |---|---|---> '''

# função para armazenar a Chave API
def store_api_key(api_key):
    f = open("api_key.txt", "w")
    f.write(api_key)
    f.close()
    
# função que lê a Chave API armazenada
def read_api_key():
    if os.path.exists("api_key.txt"):
        api_key_file = open("api_key.txt", "r")
        api_key_content = api_key_file.read()
        api_key_file.close()
    else:
        api_key_content = "" 
    
    return api_key_content

# função para armazenar a escolha de utilizar os arquivos de exemplo no "Machine Learning" como True
def store_ex_files_true():
    f = open("ex_files_opt.txt", "w")
    f.write("1")
    f.close()
    
# função para armazenar a escolha de utilizar os arquivos de exemplo no "Machine Learning" como False
def store_ex_files_false():
    f = open("ex_files_opt.txt", "w")
    f.write("0")
    f.close()

# função que lê o valor no arquivo da escolha da utilização do "Machine Larning"  
def read_ex_files_opt():
    if os.path.exists("ex_files_opt.txt"):
        with open("ex_files_opt.txt", "r") as f:
            file_content = f.read().strip()
            return bool(int(file_content))
    else:
        return False

''' <---|---|---| ARMAZENAMENTO DA CHAVE API E ESCOLHA DE USO DOS ARQUIVOS DE EXEMPLO |---|---|---> '''





''' <---|---|---| SETANDO O GOOGLE GEMINI |---|---|---> '''

# função que configura o Google Gemini
def set_gemini(api_key=""):
    try:
        # caso a Chave API esteja vazia
        if api_key == "":
            raise ErroSemChaveAPI()
        
        # setando a Google API Key
        global GOOGLE_API_KEY
        GOOGLE_API_KEY = api_key
        genai.configure(api_key=GOOGLE_API_KEY)

        # configurações de parametro
        generation_config = {
            "candidate_count": 1,
            "temperature": 0.25
        }

        # configurações de segurança
        safety_settings = {
            "HARASSMENT": "BLOCK_NONE",
            "HATE": "BLOCK_NONE",
            "SEXUAL": "BLOCK_NONE",
            "DANGEROUS": "BLOCK_NONE"
        }

        # setando o modelo generativo
        global model
        model = genai.GenerativeModel(model_name="models/gemini-1.5-pro", safety_settings=safety_settings, generation_config=generation_config)
        
     # erro caso a chave API esteja vazia
    except ErroSemChaveAPI:
        # janela de erro
        raise_janela_erro("Chave API não mencionada.", "Nenhuma Chave API foi mencionada,\ncaso não tenha nenhuma Chave API,\nacesse \"aistudio.google.com\" e\nsolicite uma Chave API nova.", "")
    
    # erro caso o limite de tokens seja ultrapassado    
    except google.api_core.exceptions.ResourceExhausted:
        # janela de erro
        raise_janela_erro("Você ultrapassou o limite de Tokens.", "Verifique se você está dentro do limite\nde Tokens do modelo. Solicite\num aumento de cota, se necessário.", "")
    
    # erro caso a chave API seja invalida
    except google.api_core.exceptions.PermissionDenied:
        # janela de erro
        raise_janela_erro("Chave API negada.", "Sua chave de API não tem as\npermissões necessárias.\nVerifique se a chave de API está\nconfigurada e tem o acesso\ncorreto.", "")
    
    # erro caso o recurso solicitado não tenha sido encontrado
    except google.api_core.exceptions.NotFound:
        # janela de erro
        raise_janela_erro("O recurso solicitado não foi encontrado.", "Verifique se todos os parâmetros\nna solicitação são válidos para\na versão da API.", "")
    
    # erro caso algo inesperado aconteça nos sistemas do google
    except google.api_core.exceptions.InternalServerError:
        # janela de erro
        raise_janela_erro("Ocorreu um erro inesperado no Google.", "Aguarde um pouco e tente enviar\nsua solicitação novamente.", "")
        
    # erro caso o serviço solicitado esteja sobrecarregado ou fora do ar
    except google.api_core.exceptions.ServiceUnavailable:
        # janela de erro
        raise_janela_erro("O serviço pode estar\ntemporariamente sobrecarregado\nou fora do ar.", "Aguarde um pouco e tente enviar\nsua solicitação novamente.", "")
    
    # lida com erros não tratados 
    except Exception as e:
        # janela de erro
        raise_janela_erro("O seguinte erro foi reportado:", "", e) 
            
''' <---|---|---| SETANDO O GOOGLE GEMINI |---|---|---> '''





''' <---|---|---| CALCULO DE TOKENS |---|---|---> '''
# seta o Google Gemini
set_gemini(read_api_key())

# procura pelo limite de Tokens do modelo
model_info = genai.get_model("models/gemini-1.5-pro")
total_tokens_input = model_info.input_token_limit
total_tokens_output = model_info.output_token_limit

# função para calcular o total de Tokens gastos no input
def calcula_tokens_input():
    try:
        # verifica se a Chave API está vazia
        if GOOGLE_API_KEY == "":
            raise ErroSemChaveAPI()
        
        if len(pdfs) != 0:
            total_tokens_input_usado = 0
            
            if bool(read_ex_files_opt()):
                total_tokens_input_usado += int(re.search(r"\d+", str(model.count_tokens(ex_files()))).group())
                
            prompt = f"Encotre as seguintes informações de cada documento:\n\n* Numero da fatura\n* Importador\n\nDocumentos:\n\n{prompt_files(pdfs)}"
            
            total_tokens_input_usado += int(re.search(r"\d+", str(model.count_tokens(prompt))).group())
            
            prompt = "Revise as informações encontradas, caso alguma informação não tenha sido encontrada, tente novamente."
            
            total_tokens_input_usado += int(re.search(r"\d+", str(model.count_tokens(prompt))).group())
            
            prompt = "Formate as informações encontradas da seguinte maneira:\n\n* O conteudo dos campos Importador deve estar em maiusculas.\n\nCaso alguma informação não tenha sido encontrada ou não foi fornecida, deixe o campo correspondente vazio.\nNão é nescessario informar quais informações não foram encontradas ou não foram fornecidas, apenas forneça o numero do pdf, o numero da fatura e o importador da seguinte forma:\n\npdf_n:\nNumero da fatura: \nImportador: \n\n(n sendo o numero do pdf)."
            
            total_tokens_input_usado += int(re.search(r"\d+", str(model.count_tokens(prompt))).group())
            
            try:
                mais_menos_simbolo = chr(177)
            except:
                mais_menos_simbolo = ""
            
            # mostra o total de Tokens utilizado
            input_token_label.configure(text=f"Tokens de Input:\t{mais_menos_simbolo} {total_tokens_input_usado} / {total_tokens_input}")
            
     # erro caso a chave API esteja vazia
    except ErroSemChaveAPI:
        
        # janela de erro
        raise_janela_erro("Chave API não mencionada.", "Nenhuma Chave API foi mencionada,\ncaso não tenha nenhuma Chave API,\nacesse \"aistudio.google.com\" e\nsolicite uma Chave API nova.", "")
    
    # erro caso o limite de tokens seja ultrapassado    
    except google.api_core.exceptions.ResourceExhausted:
        # janela de erro
        raise_janela_erro("Você ultrapassou o limite de Tokens.", "Verifique se você está dentro do limite\nde Tokens do modelo. Solicite\num aumento de cota, se necessário.", "")
    
    # erro caso a chave API seja invalida
    except google.api_core.exceptions.PermissionDenied:
        # janela de erro
        raise_janela_erro("Chave API negada.", "Sua chave de API não tem as\npermissões necessárias.\nVerifique se a chave de API está\nconfigurada e tem o acesso\ncorreto.", "")
    
    # erro caso o recurso solicitado não tenha sido encontrado
    except google.api_core.exceptions.NotFound:
        # janela de erro
        raise_janela_erro("O recurso solicitado não foi encontrado.", "Verifique se todos os parâmetros\nna solicitação são válidos para\na versão da API.", "")
    
    # erro caso algo inesperado aconteça nos sistemas do google
    except google.api_core.exceptions.InternalServerError:
        # janela de erro
        raise_janela_erro("Ocorreu um erro inesperado no Google.", "Aguarde um pouco e tente enviar\nsua solicitação novamente.", "")
        
    # erro caso o serviço solicitado esteja sobrecarregado ou fora do ar
    except google.api_core.exceptions.ServiceUnavailable:
        # janela de erro
        raise_janela_erro("O serviço pode estar\ntemporariamente sobrecarregado\nou fora do ar.", "Aguarde um pouco e tente enviar\nsua solicitação novamente.", "")
    
    # lida com erros não tratados 
    except Exception as e:
        # janela de erro
        raise_janela_erro("O seguinte erro foi reportado:", "", e) 

''' <---|---|---| CALCULO DE TOKENS |---|---|---> '''





''' <---|---|---| JANELA DE CONFIGURAÇÕES |---|---|---> '''

# função que abre a janela de configurações
def open_config_window(api_key = ""):
    # janela de configurações
    janela_config = Tk()

    # dimensões da janela
    janela_config.geometry("400x300")

    # o titulo da janela
    janela_config.title("PDF to Excel - Configurações")

    # padding e background da janela
    janela_config.config(bg="#CDDDDD", pady=30, padx=30)

    # input da API KEY
    input_api_key_label = Label(janela_config, text="Chave API:", bg="#CDDDDD", font=("Arial bold", 11), justify="left")
    input_api_key = Entry(janela_config, width=300, show="*", font=("Arial bold", 10))
    input_api_key.insert(0, api_key)
    
    input_api_key_btn = Button(janela_config, text="Atualizar a Chave API", command=lambda: store_api_key(input_api_key.get()))
    
    # radio buttons que ativa ou desativa os arquivos de exemplo
    ex_files_label = Label(janela_config, text="\"Inteligência\" da AI:", bg="#CDDDDD", font=("Arial bold", 11), justify="left")
    
    ex_files_container = Frame(janela_config, bg="#CDDDDD")
    
    ex_files_tk_var = BooleanVar(ex_files_container, read_ex_files_opt())
    
    Radiobutton(ex_files_container, text="Alto", variable=ex_files_tk_var, value=True, bg="#CDDDDD", font=("Arial bold", 11), command=lambda: [store_ex_files_true(), calcula_tokens_input()]).pack(side="left", padx=5, pady=5)
    Radiobutton(ex_files_container, text="Medio", variable=ex_files_tk_var, value=False, bg="#CDDDDD", font=("Arial bold", 11), command=lambda: [store_ex_files_false(), calcula_tokens_input()]).pack(side="left", padx=5, pady=5)
    
    ex_files_disclaimer = Label(janela_config, text="(Caso o gasto de Tokens estiver acima de sua cota é\nrecomendavel manter a \"Inteligência\" media)", bg="#CDDDDD", font=("Arial bold", 8), justify="left")

    # botão para fechar a janela de configurações
    fecha_janela_config_btn = Button(janela_config, text="Fechar", command=lambda: janela_config.destroy(), font=("Arial bold", 11))

    ''' POSICIONAMENTO DOS ELEMENTOS DA JANELA DE CONFIGURAÇÕES '''
    # setando o posicionamento dos elementos
    input_api_key_label.pack(side="top", padx=5, pady=5)
    input_api_key.pack(side="top", padx=5, pady=5)
    input_api_key_btn.pack(side="top", padx=5, pady=5)
    
    ex_files_label.pack(side="top", padx=5, pady=5)
    ex_files_container.pack(side="top", padx=5, pady=2)
    ex_files_disclaimer.pack(side="top", padx=5, pady=2)
    
    fecha_janela_config_btn.pack(side="bottom", padx=5, pady=5)
    
''' <---|---|---| JANELA DE CONFIGURAÇÕES |---|---|---> '''





''' <---|---|---| FUNÇÃO PRINCIPAL |---|---|---> '''

# numero de campos extras
num_input = 1

# dicionario contando o nome e valores dos campos extras
campos_extras = {}

# função que extrai as informações dos arquivos e gera a tabela
def sheet_generator():
    try:
        # testando a conexão
        if teste_conexao():
            # levanta um erro caso nenhum arquivo pdf tenha sido selecionado
            if len(pdfs) == 0:
                raise ErroSemArquivosPDF()
            
            # limpando o output da tabela
            output_tabela.delete(*output_tabela.get_children())
            
            # levanta um erro caso a chave API esteja vazia
            if GOOGLE_API_KEY == "":
                raise ErroSemChaveAPI()
            
            # iniciado o google gemini
            set_gemini(read_api_key())
            
            ''' INICIANDO O CHAT '''
            chat = model.start_chat(history=[])
            
            # total de token de output
            total_tokens_output_usado = 0

            # enviando os arquivos de exemplo para analise
            utilize_ex_files = bool(read_ex_files_opt())
            if utilize_ex_files:
                response = chat.send_message(ex_files())
                
                total_tokens_output_usado += int(re.search(r"\d+", str(model.count_tokens(response.text))).group())
                
            
            # prompt pedindo para que seja destacado as informações pedidas
            prompt = f"Encotre as seguintes informações de cada documento:\n\n* Numero da fatura\n* Importador\n\nDocumentos:\n\n{prompt_files(pdfs)}"

            # gerando a resposta
            response = chat.send_message(prompt)
            
            total_tokens_output_usado += int(re.search(r"\d+", str(model.count_tokens(response.text))).group())
            
            # prompt pedindo para revisar caso alguma informação não tenha sido encontrada
            prompt = "Revise as informações encontradas, caso alguma informação não tenha sido encontrada, tente novamente."
            
            # gerando a resposta
            response = chat.send_message(prompt)
            
            total_tokens_output_usado += int(re.search(r"\d+", str(model.count_tokens(response.text))).group())
            
            # prompt pedindo para formatar as informações
            prompt = "Formate as informações encontradas da seguinte maneira:\n\n* O conteudo dos campos Importador deve estar em maiusculas.\n\nCaso alguma informação não tenha sido encontrada ou não foi fornecida, deixe o campo correspondente vazio.\nNão é nescessario informar quais informações não foram encontradas ou não foram fornecidas, apenas forneça o numero do pdf, o numero da fatura e o importador da seguinte forma:\n\npdf_n:\nNumero da fatura: \nImportador: \n\n(n sendo o numero do pdf)."
            
            # gerando a resposta
            response = chat.send_message(prompt)
            
            total_tokens_output_usado += int(re.search(r"\d+", str(model.count_tokens(response.text))).group())

            # formatando e armazenando a resposta
            response_text = response.text.replace("*", "")
            
            ''' DEPURAÇÃO DA RESPOSTA DO GOOGLE GEMINI E CRIAÇÃO DA TABELA '''
            # separa as informações por arquivo e armazena em um array
            pdf_info = response_text.split("pdf_")

            # instanciando os arrays para armazenar as informações
            global num_item
            global num_fatura
            global importador
            global fronteira
            global valor
            
            num_item = [number for number in range(1, len(pdf_info))]
            num_fatura = []
            importador = []
            fronteira = []
            valor = []
            
            # verifica se o input de valor de serviço está vazio
            if input_valor.get() != "":
                input_valor_str = input_valor.get().replace("R$", "")
            else:
                input_valor_str = "0,00"
             
            if input_valor_str[-1] == ",":
                input_valor_str += "00"   
            elif input_valor_str[-2] == ",":
                input_valor_str += "0"
            elif input_valor_str[-3] != ",":
                input_valor_str += ",00"
            
            # iteração sobre cada arquivo
            for pdf in pdf_info:
                
                # retira o item vazio
                is_space = re.search(r"^\s+$", pdf) is not None
                is_pdf = re.search(r"fatura:", pdf) is not None
                
                if not is_space and is_pdf and pdf != "":
                    # procura por numero de fatura e o importador de cada arquivo
                    procura_num_fatura = re.search(r"fatura: (.+?)\n", pdf)
                    procura_importador = re.search(r"Importador: (.+?)(?:\n|$)", pdf)
                    
                    # adiciona as informações encotradas aos seus respectivos arrays
                    if procura_num_fatura:
                        if not procura_num_fatura.group(1) or procura_num_fatura.group(1).lower().strip() in ["não informado", "vazio", "não fornecido", "não especificado"] or procura_num_fatura is None:
                            num_fatura.append("")
                        else:
                            num_fatura.append(procura_num_fatura.group(1))
                    else:
                        num_fatura.append("")
                    
                    if procura_importador:
                        if not procura_importador.group(1) or procura_importador.group(1).lower().strip() in ["não informado", "vazio", "não fornecido", "não especificado"] or procura_importador is None:
                            importador.append("")
                        else:
                            importador.append(procura_importador.group(1))
                    else:
                        importador.append("")
                    
                    fronteira.append(input_fronteira.get())
                    valor.append(f"R$ {input_valor_str}")

            # soma o valor de cada fatura ao valor total 
            input_valor_format = input_valor_str.replace(".", "")
            valor_total = num_item[-1] * float(input_valor_format.replace(",", "."))
            
            # deixa uma linha vazia
            num_item.append("")
            num_fatura.append("")
            importador.append("")
            fronteira.append("")
            valor.append("")
            
            # adiciona os campos extras a tabela
            for i in range(1, len(campos_extras) // 3 + 1):
                chave_extra = campos_extras[f"input_chave_{int(i)}"].get()
                valor_extra = campos_extras[f"input_valor_{int(i)}"].get()
                
                if valor_extra != "":
                    valor_extra_str = valor_extra.replace("R$", "")
                else:
                    valor_extra_str = "0,00"
                
                if valor_extra_str[-1] == ",":
                    valor_extra_str += "00"   
                elif valor_extra_str[-2] == ",":
                    valor_extra_str += "0"
                elif valor_extra_str[-3] != ",":
                    valor_extra_str += ",00"
                
                num_item.append("")
                num_fatura.append("")
                importador.append("")
                
                fronteira.append(chave_extra)
                valor.append(f"R$ {valor_extra_str}")
                
                # soma os valores extras ao valor total
                valor_extra_format = valor_extra_str.replace(".", "")
                valor_total += float(valor_extra_format.replace(",", "."))
            
            # define a localidade para o Brasil
            locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")
            
            # formata o valor total para o formato monetario brasileiro
            valor_total_format = locale.currency(valor_total, grouping=True)
            
            # adiciona a ultima linha a tabela
            num_item.append("")
            num_fatura.append("")
            importador.append("")
            fronteira.append("TOTAL")
            valor.append(valor_total_format)
            
            # mostrando o resultado no GUI
            data_rows = list(zip(num_item, num_fatura, importador, fronteira, valor))
            
            # define a cor das linhas da tabela
            output_tabela.tag_configure("oddrow", background="#F5F5F5")
            output_tabela.tag_configure("evenrow", background="#D7D7D7")

            num_row = 1
            for row in data_rows:
                if num_row % 2 == 0:
                    output_tabela.insert("", tk.END, values=row, tags=("evenrow"))
                else:
                    output_tabela.insert("", tk.END, values=row, tags=("oddrow"))
                num_row += 1
            
            process_status.configure(text="Tabela criada")
            
            # mostrando as opções
            excel_btn.pack(side="top", padx=5, pady=5)
            refaz_btn.pack(side="top", padx=5, pady=5)
            new_sheet_btn.pack(side="top", padx=5, pady=5)
            revise_msg.pack(side="top", padx=5, pady=5)
            
            # mostra o total de tokens de output usado
            try:
                mais_menos_simbolo = chr(177)
            except:
                mais_menos_simbolo = ""
            
            output_token_label.configure(text=f"Tokens de Output:\t{mais_menos_simbolo} {total_tokens_output_usado} / {total_tokens_output}")
    
    # lida com erro caso o usuario não tenha selecionado nenhum arquivo PDF
    except ErroSemArquivosPDF:
        # janela de erro
        raise_janela_erro("Nenhum arquivo PDF foi selecionado.", "Selecione os arquivos nescessarios\npara a criação da tabela.", "")
    
    # erro caso a chave API esteja vazia
    except ErroSemChaveAPI:
        # janela de erro
        raise_janela_erro("Chave API não mencionada.", "Nenhuma Chave API foi mencionada,\ncaso não tenha nenhuma Chave API,\nacesse \"aistudio.google.com\" e\nsolicite uma Chave API nova.", "")
    
    # erro caso o limite de tokens seja ultrapassado    
    except google.api_core.exceptions.ResourceExhausted:
        # janela de erro
        raise_janela_erro("Você ultrapassou o limite de Tokens.", "Verifique se você está dentro do limite\nde Tokens do modelo. Solicite\num aumento de cota, se necessário.", "")
    
    # erro caso a chave API seja invalida
    except google.api_core.exceptions.PermissionDenied:
        # janela de erro
        raise_janela_erro("Chave API negada.", "Sua chave de API não tem as\npermissões necessárias.\nVerifique se a chave de API está\nconfigurada e tem o acesso\ncorreto.", "")
    
    # erro caso o recurso solicitado não tenha sido encontrado
    except google.api_core.exceptions.NotFound:
        # janela de erro
        raise_janela_erro("O recurso solicitado não foi encontrado.", "Verifique se todos os parâmetros\nna solicitação são válidos para\na versão da API.", "")
    
    # erro caso algo inesperado aconteça nos sistemas do google
    except google.api_core.exceptions.InternalServerError:
        # janela de erro
        raise_janela_erro("Ocorreu um erro inesperado no Google.", "Aguarde um pouco e tente enviar\nsua solicitação novamente.", "")
        
    # erro caso o serviço solicitado esteja sobrecarregado ou fora do ar
    except google.api_core.exceptions.ServiceUnavailable:
        # janela de erro
        raise_janela_erro("O serviço pode estar\ntemporariamente sobrecarregado\nou fora do ar.", "Aguarde um pouco e tente enviar\nsua solicitação novamente.", "")
    
    # lida com erros não tratados 
    except Exception as e:
        # janela de erro
        raise_janela_erro("O seguinte erro foi reportado:", "", e) 
         
''' <---|---|---| FUNÇÃO PRINCIPAL |---|---|---> '''




''' <---|---|---| TO EXCEL |---|---|---> '''

# função que cria um arquivo Excel
def to_excel():
   # Cria um novo workbook
    wb = openpyxl.Workbook()

    # Cria a planilha ativa
    ws = wb.active

    # Define o título da planilha
    ws.title = "Detalhamento"
    
    ''' ESTILO DA PLANILHA '''
    # fonte principal
    fonte_negrito = Font(name="Calibri", size=11, bold=True)
    
    # fonte de destaque
    title_font = Font(name="Elephant", size=12, bold=True, color="548ED4")
    
    # fonte do valor total
    total_font = Font(name="Calibri", size=11, bold=True, color="FF0000")

    # alinhamento
    alinhamento_centro = Alignment(horizontal="center")

    # borda
    borda = Border(left=Side(border_style="medium", color="000000"),
                    right=Side(border_style="medium", color="000000"),
                    top=Side(border_style="medium", color="000000"),
                    bottom=Side(border_style="medium", color="000000"))

    # preenchimento
    preenchimento = PatternFill(start_color="FFFF00",
                    end_color="FFFF00",
                    fill_type="solid")
    
    ''' CABEÇALHO DE INFORMAÇÕES '''
    ws.cell(row=3, column=2).value = "COSTA & SEABRA  ASSESSORIA ADUANEIRA  LTDA"
    ws.cell(row=3, column=2).font = title_font

    ws.cell(row=4, column=2).value = "( CNPJ : 17.572.568/0001-80 )"
    ws.cell(row=4, column=2).font = fonte_negrito

    ws.cell(row=5, column=2).value = "RUA DEMÉTRIO RIBEIRO ,Nº 365  BAIRRO - CARVALHO"
    ws.cell(row=5, column=2).font = title_font

    ws.cell(row=6, column=2).value = "JAGUARÃO – RS - BRASIL    CEP 96300-000"
    ws.cell(row=6, column=2).font = title_font
    
    # data atual
    hoje = datetime.date.today()

    # separa o dia, mes e ano
    dia = hoje.day
    mes = hoje.month
    ano = hoje.year

    # seleciona o nome do mes
    if int(mes) == 1:
        mes_str = "JANEIRO"
    elif int(mes) == 2:
        mes_str = "FEVEREIRO"
    elif int(mes) == 3:
        mes_str = "MARÇO"
    elif int(mes) == 4:
        mes_str = "ABRIL"
    elif int(mes) == 5:
        mes_str = "MAIO"
    elif int(mes) == 6:
        mes_str = "JUNHO"
    elif int(mes) == 7:
        mes_str = "JULHO"
    elif int(mes) == 8:
        mes_str = "AGOSTO"
    elif int(mes) == 9:
        mes_str = "SETEMBRO"
    elif int(mes) == 10:
        mes_str = "OUTUBRO"
    elif int(mes) == 11:
        mes_str = "NOVEMBRO"
    elif int(mes) == 12:
        mes_str = "DEZEMBRO"
        
    hoje_str = f"JAGUARÃO, {dia} DE {mes_str} DE {ano}"
    ws.cell(row=8, column=2).value = hoje_str
    ws.cell(row=8, column=2).font = fonte_negrito

    ws.cell(row=10, column=2).value = "ESTAMOS LHE ENVIANDO ABAIXO OS DADOS REFERENTES AO DESEMBARAÇO"
    ws.cell(row=10, column=2).font = fonte_negrito

    ws.cell(row=11, column=2).value = "ADUANEIRO REALIZADO POR NOSSO ESCRITÓRIO EM FRONTEIRA:"
    ws.cell(row=11, column=2).font = fonte_negrito

    detalhamento_str = f"DETALHAMENTO DA FATURA {detalhamento_input.get()}"

    ws.cell(row=13, column=2).value = detalhamento_str
    ws.cell(row=13, column=2).font = fonte_negrito
    
    ''' TABELA '''
    # Define os nomes das colunas
    column_names = ["ITEM", "FATURA", "IMPORTADOR", "FRONTEIRA", "VALOR"]
    
    # Adiciona o nome das colunas a tabela
    for i, name in enumerate(column_names):
        ws.cell(row=15, column=i + 1).value = name
        ws.cell(row=15, column=i + 1).font = fonte_negrito
        ws.cell(row=15, column=i + 1).alignment = alinhamento_centro
        ws.cell(row=15, column=i + 1).border = borda
        ws.cell(row=15, column=i + 1).fill = preenchimento
        
    # linha inicial da tabela
    linha = 16
    
    # adiciona os itens a tabela
    for item in num_item:
        # Escreve os valores de cada coluna na linha atual
        ws.cell(row=linha, column=1).value = item
        ws.cell(row=linha, column=2).value = num_fatura[linha - 16]
        ws.cell(row=linha, column=3).value = importador[linha - 16]
        ws.cell(row=linha, column=4).value = fronteira[linha - 16]
        ws.cell(row=linha, column=5).value = valor[linha - 16]
        
        ws.cell(row=linha, column=1).border = borda
        ws.cell(row=linha, column=2).border = borda
        ws.cell(row=linha, column=3).border = borda
        ws.cell(row=linha, column=4).border = borda
        ws.cell(row=linha, column=5).border = borda
        
        ws.cell(row=linha, column=1).alignment = alinhamento_centro
        ws.cell(row=linha, column=2).alignment = alinhamento_centro
        ws.cell(row=linha, column=3).alignment = alinhamento_centro
        ws.cell(row=linha, column=4).alignment = alinhamento_centro
        ws.cell(row=linha, column=5).alignment = alinhamento_centro
        
        if fronteira[linha - 16] == "TOTAL":
            ws.cell(row=linha, column=4).fill = preenchimento
            ws.cell(row=linha, column=5).font = total_font

        # Incrementa o contador da linha para a próxima linha
        linha += 1
    
    # adicionando a linha de assinatura
    linha_assin = linha + 1
    nome_assin = linha + 2
    
    ws.cell(row=linha_assin, column=3).value = "___________________________________"
    ws.cell(row=nome_assin, column=3).value = "JOICE H C SEABRA"
    
    ws.cell(row=linha_assin, column=3).font = fonte_negrito
    ws.cell(row=nome_assin, column=3).font = fonte_negrito
    
    ws.cell(row=linha_assin, column=3).alignment = alinhamento_centro
    ws.cell(row=nome_assin, column=3).alignment = alinhamento_centro
    
    
    ''' SALVANDO A PLANILHA '''
    # seletor do local de salvamento do arquivo
    save_location = filedialog.asksaveasfile(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    
    if save_location:
        # Salva a planilha
        wb.save(save_location.name)
        
        if len(save_location.name) >= 20:
            save_location_name = f"{save_location.name[:21]}..."
        else:
            save_location_name = save_location.name
        
        process_status.configure(text=f"Arquivo Excel salvo em:\n{save_location_name}")

''' <---|---|---| TO EXCEL |---|---|---> '''





''' <---|---|---| OUTRAS FUNÇÕES |---|---|---> '''

# função para gerar uma nova planilha
def new_sheet():  
    global num_input
    output_tabela.delete(*output_tabela.get_children())
    
    quant_arquivos.configure(text="Selecione os arquivos PDF")
    nome_arquivos.configure(text="")
    
    detalhamento_input.delete(0, tk.END)
    input_fronteira.delete(0, tk.END)
    input_valor.delete(0, tk.END)
    
    input_token_label.configure(text=f"Tokens de Input:\t0 / {total_tokens_input}")
    output_token_label.configure(text=f"Tokens de Output:\t0 / {total_tokens_output}")
    
    if campos_extras:
        for key, value in campos_extras.items():
            if "input_container_" in key:
                value.destroy()

        campos_extras.clear()
        
    num_input = 1
    
    troca_arquivos()
    
''' <---|---|---| OUTRAS FUNÇÕES |---|---|---> '''




   
''' <---|---|---| GUI |---|---|---> '''

# função para adicionar os input dos campos extras
def adicionar_campo(extra_input_container): 
    global num_input
    # Criar novos nomes de campo com base no número atual
    novo_nome_container = f"input_container_{num_input}"
    novo_nome_chave = f"input_chave_{num_input}"
    novo_nome_label = f"label_input_{num_input}"
    novo_nome_valor = f"input_valor_{num_input}"

    # Criar campos de entrada
    novo_input_container = Frame(extra_input_container, name=novo_nome_container, bg="#CDDDDD")
    novo_campo_chave = Entry(novo_input_container, name=novo_nome_chave, width=20)
    novo_label = Label(novo_input_container, name=novo_nome_label, text="-->\tR$", bg="#CDDDDD")
    novo_campo_valor = Entry(novo_input_container, name=novo_nome_valor, width=15, justify="right")

    # Posicionar os campos na interface
    novo_input_container.pack(side="top", padx=5, pady=5)
    novo_campo_chave.pack(side="left", padx=5, pady=5)
    novo_label.pack(side="left", padx=2, pady=5)
    novo_campo_valor.pack(side="left", padx=2, pady=5)
    
    novos_campos = {
        novo_nome_container: novo_input_container,
        novo_nome_chave: novo_campo_chave,
        novo_nome_valor: novo_campo_valor
    }
    
    campos_extras.update(novos_campos)
    
    num_input += 1
    
''' JANELA PRINCIPAL '''
# janela
janela = Tk()

# dimensões da janela
janela.geometry("1400x600")

# o titulo da janela
janela.title("PDF to Excel - Detalhamentos")

# padding e background da janela
janela.config(bg="#CDDDDD", pady=50, padx=30)

''' LADO ESQUERDO DA JANELA '''
# o container esquerdo
left_container = Frame(janela, bg="#CDDDDD")

# label que mostra a quantidade de arquivos selecionados
quant_arquivos = Label(left_container, text="Selecione os arquivos PDF", bg="#CDDDDD", font=("Arial bold", 12))

# label que mostra o caminho dos arquivos selecionados
nome_arquivos = Label(left_container, text="", fg="blue", bg="#CDDDDD", font=("Arial bold", 10))

# o container dos botões de procura de arquivo
procura_arquivo_container = Frame(left_container, bg="#CDDDDD")

# o botão de busca de arquivos
procura_arquivo_btn = Button(procura_arquivo_container, text="Selecionar arquivos", command=lambda: [procura_arquivos(), calcula_tokens_input()], bg="#A599B5", width=15)

# o botão para trocar de arquivos
troca_arquivo_btn = Button(procura_arquivo_container, text="Escolher outros arquivos", command=lambda: [troca_arquivos(), calcula_tokens_input()], bg="#A64B81", width=20)

# o input do "nome" do detalhamento
detalhamento_label = Label(left_container, text="Detalhamento da Fatura:", bg="#CDDDDD", font=("Arial bold", 10))
detalhamento_input = Entry(left_container, width=40)

# o container dos inputs de fronteira e valor
fronteira_valor_container = Frame(left_container, bg="#CDDDDD")

# o input de fronteira
input_fronteira_label = Label(fronteira_valor_container, text="Fronteira:", bg="#CDDDDD", font=("Arial bold", 10))
input_fronteira = Entry(fronteira_valor_container, justify="right", width=15)

# o input do valor
input_valor_label = Label(fronteira_valor_container, text="Valor de serviço:\tR$", bg="#CDDDDD", font=("Arial bold", 10))
input_valor = Entry(fronteira_valor_container, justify="right", width=15)

# o container dos campos extras
extra_input_container = Frame(left_container, bg="#CDDDDD")

# botão para adicionar campos extras
extra_input_btn = Button(left_container, text="Adicionar um campo extra", command=lambda: adicionar_campo(extra_input_container), bg="#9667f5")

# o botão que inicia a criação da tabela
inicia_btn = Button(left_container, text="Criar Tabela", command=sheet_generator, bg="#0EB1D2", width=15)

''' CENTRO DA JANELA '''
# o resultado da tabela
# nome das colunas
column_names = ["ITEM", "FATURA", "IMPORTADOR", "FRONTEIRA", "VALOR"]
output_tabela = ttk.Treeview(janela, columns=column_names, show="headings")
output_tabela.config(height=20)
for column in column_names:
    output_tabela.heading(column, text=column)
    
# configuração da largura das colunas
output_tabela.column("ITEM", width=50)
output_tabela.column("FATURA", width=150)
output_tabela.column("IMPORTADOR", width=300)
output_tabela.column("FRONTEIRA", width=100)
output_tabela.column("VALOR", width=100)

''' LADO DIREITO DA JANELA '''
# o container direito
right_container = Frame(janela, bg="#CDDDDD")

# uma label que informa se atabela foi criada
process_status = Label(right_container, text="", bg="#CDDDDD", font=("Arial bold", 12))

# o botão parar baixar o excel
excel_btn = Button(right_container, text="Gerar Excel", command=to_excel, bg="#388659", width=15)

# o botão para refazer a tabela
refaz_btn = Button(right_container, text="Refazer Tabela", command=sheet_generator, bg="#F4FF52", width=15)

# o botão para resetar o programa
new_sheet_btn = Button(right_container, text="Criar Nova Tabela", command=new_sheet, bg="#DD6E42", width=15)

# uma label dizendo que é recomendavel que as informações sejam revisadas
revise_msg = Label(right_container, text="É altamente recomendado que\nas informações sejam revisadas.\nCaso as informações estejam\nmuito incondizentes, é recomendado\nque a tabela seja refeita.", bg="#CDDDDD", font=("Arial bold", 10), justify="left")


''' BOTÃO DE CONFIGURAÇÃO '''
# o botão de configuração da API KEY
config_btn_image = Image.open("settings_icon.jpg")
config_btn_image_resize = config_btn_image.resize((30, 30))
config_btn_photo = ImageTk.PhotoImage(config_btn_image_resize)


config_btn = Button(janela, image=config_btn_photo, width=30, height=30, bg="#B3B3B3", command=lambda: open_config_window(GOOGLE_API_KEY))
config_btn.image = config_btn_photo

''' OUTPUT DE TOKENS '''
# informação de quantos tokens já foram utilizados
token_label_container = Frame(janela, bg="#CDDDDD")
input_token_label = Label(token_label_container, text=f"Tokens de Input:\t0 / {total_tokens_input}", bg="#CDDDDD", font=("Arial bold", 9))
output_token_label = Label(token_label_container, text=f"Tokens de Output:\t0 / {total_tokens_output}", bg="#CDDDDD", font=("Arial bold", 9))

# label mostrano o nome do modelo generativo
model_name_label = Label(janela, text="Feito com Google Gemini 1.5-pro", bg="#CDDDDD", font=("Arial bold", 8), justify="left")

''' POSICIONAMENTO DOS ELEMENTOS DA JANELA PRINCIPAL'''
# setando o posicionamento dos elementos
# lado esquerdo
left_container.grid(row=0, column=0, padx=10, pady=5, sticky=N)

quant_arquivos.pack(side="top", padx=5, pady=5)
nome_arquivos.pack(side="top", padx=5, pady=5)

procura_arquivo_container.pack(side="top", padx=5, pady=5)
procura_arquivo_btn.pack(side="top", padx=5, pady=5)

detalhamento_label.pack(side="top", padx=5, pady=5)
detalhamento_input.pack(side="top", padx=5, pady=5)

fronteira_valor_container.pack(side="top", padx=5, pady=5)
input_fronteira_label.grid(column=0, row=0, padx=5, pady=5)
input_fronteira.grid(column=1, row=0, padx=5, pady=5)
input_valor_label.grid(column=0, row=1, padx=2, pady=5)
input_valor.grid(column=1, row=1, padx=2, pady=5)

extra_input_container.pack(side="top", padx=5, pady=5)
extra_input_btn.pack(side="top", padx=5, pady=5)

inicia_btn.pack(side="top", padx=5, pady=5)

# centro
output_tabela.grid(row=0, column=1, padx=10, pady=5, sticky=N)

# lado direito
right_container.grid(row=0, column=2, padx=10, pady=5, sticky=N)

process_status.pack(side="top", padx=5, pady=5)

# botão de configuração
config_btn.grid(row=0, column=3, padx=5, pady=5, sticky=NW)

# output de tokens
token_label_container.grid(row=1, column=1, padx=5, pady=5, sticky=N)
input_token_label.pack(side="left", padx=5, pady=5)
output_token_label.pack(side="left", padx=5, pady=5) 

# label do modelo gnerativo
model_name_label.grid(row=1, column=0, padx=5, pady=5, sticky=NW)

# abre a janela principal
janela.mainloop()

''' <---|---|---| GUI |---|---|---> '''

''' Kauã Costa Seabra - 05/2024'''